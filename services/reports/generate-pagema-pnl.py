#!/usr/bin/env python3
"""
Generate Pagema Ltd P&L by Branch reports in HTML, PDF, MD, and XLSX formats.
Reads raw JSON from pagema-pnl-raw.json (fetched from Xero API).
"""
import json
import os
from collections import OrderedDict

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
RAW_FILE = os.path.join(OUTPUT_DIR, 'pagema-pnl-raw.json')
BASE_NAME = 'pagema-pnl-by-branch-dec25-feb26'

BRANCHES = ['Brent', 'Chiswick', 'Peckham', 'Shoreditch', 'Wandsworth']
MONTHS = ['December 2025', 'January 2026', 'February 2026']

# Brand colors
PRIMARY = '#1B365D'
SECONDARY = '#D4A84B'
GREEN = '#2E7D32'
RED = '#C62828'
LIGHT_BG = '#F5F5F5'
WHITE = '#FFFFFF'

# ─── Parse Xero Report ───

def parse_report(report_json):
    """Parse a Xero P&L report into an ordered dict of {label: amount}.
    Also returns section_map: {label: section_title} for ordering."""
    report = report_json['Reports'][0]
    items = OrderedDict()
    section_map = {}
    for section in report['Rows']:
        if section['RowType'] == 'Header':
            continue
        title = section.get('Title', '')
        if section.get('Rows'):
            for row in section['Rows']:
                label = row['Cells'][0]['Value']
                val_str = row['Cells'][1]['Value'] if len(row['Cells']) > 1 else ''
                amount = float(val_str) if val_str else 0.0
                if row['RowType'] == 'SummaryRow':
                    key = f'__TOTAL__{label}'
                    items[key] = amount
                    section_map[key] = title
                elif title:
                    items[label] = amount
                    section_map[label] = title
                else:
                    # Gross Profit or Net Profit standalone rows
                    key = f'__KEY__{label}'
                    items[key] = amount
                    section_map[key] = '__standalone__'
    return items, section_map


def build_all_data(raw):
    """Build structured data: {month: {branch/Total: {label: amount}}}
    Also builds a global section_map for proper ordering."""
    all_data = {}
    global_section_map = {}
    for month in MONTHS:
        all_data[month] = {}
        for key in ['Total'] + BRANCHES:
            items, section_map = parse_report(raw[month][key])
            all_data[month][key] = items
            global_section_map.update(section_map)
    return all_data, global_section_map


def get_all_labels(all_data, section_map):
    """Get a unified ordered list of labels, respecting section ordering.
    Xero P&L sections always follow: Income -> Cost of Sales -> Gross Profit ->
    Other Income -> Operating Expenses -> Net Profit."""

    SECTION_ORDER = [
        'Income',
        '__total_income__',       # Total Income
        'Less Cost of Sales',
        '__total_cos__',          # Total Cost of Sales
        '__gross_profit__',       # Gross Profit
        'Plus Other Income',
        '__total_other__',        # Total Other Income
        'Less Operating Expenses',
        '__total_opex__',         # Total Operating Expenses
        '__net_profit__',         # Net Profit
    ]

    # Collect all labels per section
    section_labels = OrderedDict()
    for s in SECTION_ORDER:
        section_labels[s] = []

    # Gather all labels from all reports
    all_label_set = set()
    for month in MONTHS:
        for key in ['Total'] + BRANCHES:
            for label in all_data[month][key]:
                all_label_set.add(label)

    for label in all_label_set:
        sec = section_map.get(label, '')
        is_total = label.startswith('__TOTAL__')
        is_key = label.startswith('__KEY__')

        if is_key:
            disp = display_label(label)
            if 'Gross' in disp:
                section_labels['__gross_profit__'].append(label)
            elif 'Net' in disp:
                section_labels['__net_profit__'].append(label)
            else:
                section_labels['__net_profit__'].append(label)
        elif is_total:
            disp = display_label(label)
            if 'Income' in disp and 'Other' not in disp:
                section_labels['__total_income__'].append(label)
            elif 'Cost of Sales' in disp:
                section_labels['__total_cos__'].append(label)
            elif 'Other' in disp:
                section_labels['__total_other__'].append(label)
            elif 'Operating' in disp or 'Expense' in disp:
                section_labels['__total_opex__'].append(label)
            else:
                section_labels['__net_profit__'].append(label)
        elif sec == 'Income':
            section_labels['Income'].append(label)
        elif sec == 'Less Cost of Sales':
            section_labels['Less Cost of Sales'].append(label)
        elif sec == 'Plus Other Income':
            section_labels['Plus Other Income'].append(label)
        elif sec == 'Less Operating Expenses':
            section_labels['Less Operating Expenses'].append(label)
        else:
            # Fallback - put at end
            section_labels['__net_profit__'].append(label)

    # Sort items within each section alphabetically for consistency
    for key in section_labels:
        if not key.startswith('__'):
            section_labels[key].sort(key=lambda x: display_label(x).lower())

    # Flatten
    result = []
    for key in SECTION_ORDER:
        result.extend(section_labels[key])

    return result


def compute_unassigned(all_data):
    """Compute Unassigned = Total - sum(branches) for each line item."""
    result = {}
    for month in MONTHS:
        result[month] = {}
        total_data = all_data[month]['Total']
        for label in total_data:
            branch_sum = sum(all_data[month].get(b, {}).get(label, 0.0) for b in BRANCHES)
            result[month][label] = total_data[label] - branch_sum
    return result


def fmt_gbp(val):
    """Format as GBP with commas. Negative in parentheses."""
    if val == 0:
        return '-'
    if val < 0:
        return f'({chr(163)}{abs(val):,.2f})'
    return f'{chr(163)}{val:,.2f}'


def fmt_pct(val, total_income):
    """Format as percentage of total income."""
    if total_income == 0 or val == 0:
        return '-'
    pct = (val / total_income) * 100
    if pct < 0:
        return f'({abs(pct):.1f}%)'
    return f'{pct:.1f}%'


def display_label(label):
    """Clean up internal label markers for display."""
    if label.startswith('__TOTAL__'):
        return label.replace('__TOTAL__', '')
    if label.startswith('__KEY__'):
        return label.replace('__KEY__', '')
    return label


def is_total_row(label):
    return label.startswith('__TOTAL__') or label.startswith('__KEY__')


def get_total_income(data_dict):
    """Get total income for a branch/column."""
    return data_dict.get('__TOTAL__Total Income', 0.0)


# ─── Prepare column data for a month ───

def month_table_data(all_data, unassigned, month, section_map):
    """Return (labels, columns) where columns = {colname: {label: amount}}"""
    labels = get_all_labels(all_data, section_map)
    columns = OrderedDict()
    for b in BRANCHES:
        columns[b] = all_data[month].get(b, {})
    columns['Unassigned'] = unassigned[month]
    columns['Total'] = all_data[month]['Total']
    return labels, columns


# ─── Section grouping ───

def group_labels_by_section(labels):
    """Group labels into sections based on Total/Key markers."""
    sections = []
    current_items = []
    for label in labels:
        current_items.append(label)
        if is_total_row(label):
            sections.append(current_items)
            current_items = []
    if current_items:
        sections.append(current_items)
    return sections


# ═══════════════════════════════════════════════════════
# HTML Generation
# ═══════════════════════════════════════════════════════

def generate_html(all_data, unassigned, section_map):
    col_headers = BRANCHES + ['Unassigned', 'Total']

    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Pagema Ltd &mdash; P&amp;L by Branch</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Segoe UI', Arial, sans-serif; background: {LIGHT_BG}; color: #333; padding: 20px; }}
  .container {{ max-width: 1400px; margin: 0 auto; }}
  h1 {{ color: {PRIMARY}; font-size: 24px; margin-bottom: 4px; }}
  h2 {{ color: {PRIMARY}; font-size: 18px; margin: 30px 0 10px; border-bottom: 2px solid {SECONDARY}; padding-bottom: 4px; }}
  .subtitle {{ color: #666; font-size: 14px; margin-bottom: 20px; }}
  table {{ width: 100%; border-collapse: collapse; margin-bottom: 30px; font-size: 12px; }}
  th {{ background: {PRIMARY}; color: white; padding: 8px 6px; text-align: right; font-weight: 600; white-space: nowrap; }}
  th:first-child {{ text-align: left; min-width: 200px; }}
  td {{ padding: 5px 6px; text-align: right; border-bottom: 1px solid #ddd; white-space: nowrap; }}
  td:first-child {{ text-align: left; font-weight: 400; }}
  tr:nth-child(even) {{ background: #f9f9f9; }}
  tr.total-row {{ font-weight: 700; background: #e8e8e8 !important; border-top: 2px solid {PRIMARY}; }}
  tr.key-row {{ font-weight: 700; background: #dce4ed !important; border-top: 2px solid {SECONDARY}; border-bottom: 2px solid {SECONDARY}; }}
  .negative {{ color: {RED}; }}
  .positive {{ color: {GREEN}; }}
  .pct {{ color: #888; font-size: 10px; }}
  .col-pair {{ text-align: right; }}
  .flag {{ background: #fff3cd; }}
  .report-footer {{ text-align: center; color: #999; font-size: 11px; margin-top: 40px; padding-top: 10px; border-top: 1px solid #ddd; }}
</style>
</head>
<body>
<div class="container">
<h1>Pagema Ltd &mdash; P&amp;L (No Shareholder) by Branch</h1>
<p class="subtitle">December 2025 &ndash; February 2026 | Generated {get_today()}</p>
'''

    # Monthly tables
    for month in MONTHS:
        labels, columns = month_table_data(all_data, unassigned, month, section_map)
        html += f'<h2>{month}</h2>\n'
        html += render_html_table(labels, columns, col_headers)

    # Summary table
    html += '<h2>3-Month Summary (Dec 2025 &ndash; Feb 2026)</h2>\n'
    summary_labels, summary_columns = build_summary(all_data, unassigned, section_map)
    html += render_html_table(summary_labels, summary_columns, col_headers)

    html += '<div class="report-footer">Pagema Ltd &mdash; Confidential Financial Report</div>\n'
    html += '</div></body></html>'

    path = os.path.join(OUTPUT_DIR, f'{BASE_NAME}.html')
    with open(path, 'w') as f:
        f.write(html)
    print(f'HTML: {path}')


def render_html_table(labels, columns, col_headers):
    """Render an HTML table for one month or summary."""
    html = '<table>\n<thead><tr><th>Line Item</th>'
    for col in col_headers:
        html += f'<th>{col}<br><span class="pct">&pound; / %</span></th>'
    html += '</tr></thead>\n<tbody>\n'

    sections = group_labels_by_section(labels)

    for section_labels in sections:
        for label in section_labels:
            is_total = label.startswith('__TOTAL__')
            is_key = label.startswith('__KEY__')
            row_class = ''
            if is_total:
                row_class = ' class="total-row"'
            elif is_key:
                row_class = ' class="key-row"'

            disp = display_label(label)
            html += f'<tr{row_class}><td>{disp}</td>'

            for col in col_headers:
                val = columns.get(col, {}).get(label, 0.0)
                ti = get_total_income(columns.get(col, {}))
                gbp = fmt_gbp(val)
                pct = fmt_pct(val, ti)

                val_class = ''
                if val < 0:
                    val_class = ' class="negative"'
                elif val > 0 and (is_key or is_total) and ('Profit' in disp):
                    val_class = ' class="positive"'

                # Flag significant unassigned
                flag = ''
                if col == 'Unassigned' and abs(val) > 0:
                    total_val = columns.get('Total', {}).get(label, 0.0)
                    if total_val != 0 and abs(val / total_val) > 0.05:
                        flag = ' class="flag"'

                html += f'<td{val_class}{flag}>{gbp} <span class="pct">{pct}</span></td>'
            html += '</tr>\n'

    html += '</tbody></table>\n'
    return html


def build_summary(all_data, unassigned, section_map):
    """Build 3-month summary by summing across all months."""
    labels = get_all_labels(all_data, section_map)
    col_keys = BRANCHES + ['Unassigned', 'Total']
    summary = OrderedDict()

    for col in BRANCHES:
        summary[col] = {}
        for label in labels:
            summary[col][label] = sum(all_data[m].get(col, {}).get(label, 0.0) for m in MONTHS)

    summary['Unassigned'] = {}
    for label in labels:
        summary['Unassigned'][label] = sum(unassigned[m].get(label, 0.0) for m in MONTHS)

    summary['Total'] = {}
    for label in labels:
        summary['Total'][label] = sum(all_data[m]['Total'].get(label, 0.0) for m in MONTHS)

    return labels, summary


def get_today():
    from datetime import date
    return date.today().strftime('%d %B %Y')


# ═══════════════════════════════════════════════════════
# Markdown Generation
# ═══════════════════════════════════════════════════════

def generate_md(all_data, unassigned, section_map):
    col_headers = BRANCHES + ['Unassigned', 'Total']
    lines = []
    lines.append(f'# Pagema Ltd -- P&L (No Shareholder) by Branch')
    lines.append(f'**December 2025 -- February 2026** | Generated {get_today()}')
    lines.append('')

    for month in MONTHS:
        labels, columns = month_table_data(all_data, unassigned, month, section_map)
        lines.append(f'## {month}')
        lines.append('')
        lines += render_md_table(labels, columns, col_headers)
        lines.append('')

    lines.append('## 3-Month Summary (Dec 2025 -- Feb 2026)')
    lines.append('')
    summary_labels, summary_columns = build_summary(all_data, unassigned, section_map)
    lines += render_md_table(summary_labels, summary_columns, col_headers)
    lines.append('')
    lines.append('---')
    lines.append('*Pagema Ltd -- Confidential Financial Report*')

    path = os.path.join(OUTPUT_DIR, f'{BASE_NAME}.md')
    with open(path, 'w') as f:
        f.write('\n'.join(lines))
    print(f'MD: {path}')


def render_md_table(labels, columns, col_headers):
    """Render a Markdown table."""
    lines = []
    header = '| Line Item | ' + ' | '.join(col_headers) + ' |'
    sep = '|---|' + '|'.join(['---:'] * len(col_headers)) + '|'
    lines.append(header)
    lines.append(sep)

    for label in labels:
        disp = display_label(label)
        if is_total_row(label):
            disp = f'**{disp}**'
        row = f'| {disp} |'
        for col in col_headers:
            val = columns.get(col, {}).get(label, 0.0)
            row += f' {fmt_gbp(val)} |'
        lines.append(row)

    return lines


# ═══════════════════════════════════════════════════════
# Excel Generation
# ═══════════════════════════════════════════════════════

def generate_xlsx(all_data, unassigned, section_map):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

    wb = Workbook()
    col_headers = BRANCHES + ['Unassigned', 'Total']

    header_fill = PatternFill(start_color='1B365D', end_color='1B365D', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    total_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
    key_fill = PatternFill(start_color='DCE4ED', end_color='DCE4ED', fill_type='solid')
    bold_font = Font(bold=True, size=10)
    normal_font = Font(size=10)
    red_font = Font(color='C62828', size=10)
    red_bold = Font(color='C62828', bold=True, size=10)
    green_bold = Font(color='2E7D32', bold=True, size=10)
    thin_border = Border(bottom=Side(style='thin', color='DDDDDD'))
    gbp_fmt = '#,##0.00;(#,##0.00);"-"'
    pct_fmt = '0.0%;(0.0%);"-"'

    sheets_data = []
    for month in MONTHS:
        labels, columns = month_table_data(all_data, unassigned, month, section_map)
        sheets_data.append((month, labels, columns))

    summary_labels, summary_columns = build_summary(all_data, unassigned, section_map)
    sheets_data.append(('Summary', summary_labels, summary_columns))

    for idx, (sheet_name, labels, columns) in enumerate(sheets_data):
        if idx == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + len(col_headers) * 2)
        ws.cell(1, 1, f'Pagema Ltd -- P&L by Branch: {sheet_name}').font = Font(bold=True, size=14, color='1B365D')

        # Header row
        row_num = 3
        ws.cell(row_num, 1, 'Line Item').font = header_font
        ws.cell(row_num, 1).fill = header_fill
        for i, col in enumerate(col_headers):
            c_gbp = 2 + i * 2
            c_pct = 3 + i * 2
            ws.cell(row_num, c_gbp, f'{col} (GBP)').font = header_font
            ws.cell(row_num, c_gbp).fill = header_fill
            ws.cell(row_num, c_gbp).alignment = Alignment(horizontal='right')
            ws.cell(row_num, c_pct, f'{col} (%)').font = header_font
            ws.cell(row_num, c_pct).fill = header_fill
            ws.cell(row_num, c_pct).alignment = Alignment(horizontal='right')

        # Data rows
        row_num = 4
        for label in labels:
            is_tot = label.startswith('__TOTAL__')
            is_k = label.startswith('__KEY__')
            disp = display_label(label)

            ws.cell(row_num, 1, disp)
            if is_tot or is_k:
                ws.cell(row_num, 1).font = bold_font
                for c in range(1, 2 + len(col_headers) * 2):
                    ws.cell(row_num, c).fill = key_fill if is_k else total_fill
            else:
                ws.cell(row_num, 1).font = normal_font

            for i, col in enumerate(col_headers):
                val = columns.get(col, {}).get(label, 0.0)
                ti = get_total_income(columns.get(col, {}))

                c_gbp = 2 + i * 2
                c_pct = 3 + i * 2

                cell_gbp = ws.cell(row_num, c_gbp, val)
                cell_gbp.number_format = gbp_fmt
                cell_gbp.alignment = Alignment(horizontal='right')

                if ti != 0:
                    pct_val = val / ti
                else:
                    pct_val = 0
                cell_pct = ws.cell(row_num, c_pct, pct_val)
                cell_pct.number_format = pct_fmt
                cell_pct.alignment = Alignment(horizontal='right')

                # Font coloring
                if val < 0:
                    font = red_bold if (is_tot or is_k) else red_font
                    cell_gbp.font = font
                    cell_pct.font = font
                elif (is_tot or is_k):
                    if 'Profit' in disp and val > 0:
                        cell_gbp.font = green_bold
                        cell_pct.font = green_bold
                    else:
                        cell_gbp.font = bold_font
                        cell_pct.font = bold_font
                else:
                    cell_gbp.font = normal_font
                    cell_pct.font = normal_font

                cell_gbp.border = thin_border
                cell_pct.border = thin_border

            row_num += 1

        # Auto-width
        from openpyxl.utils import get_column_letter
        ws.column_dimensions['A'].width = 35
        for i in range(len(col_headers)):
            ws.column_dimensions[get_column_letter(2 + i * 2)].width = 16
            ws.column_dimensions[get_column_letter(3 + i * 2)].width = 10

    path = os.path.join(OUTPUT_DIR, f'{BASE_NAME}.xlsx')
    wb.save(path)
    print(f'XLSX: {path}')


# ═══════════════════════════════════════════════════════
# PDF Generation
# ═══════════════════════════════════════════════════════

def generate_pdf(all_data, unassigned, section_map):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle

    path = os.path.join(OUTPUT_DIR, f'{BASE_NAME}.pdf')
    col_headers = BRANCHES + ['Unassigned', 'Total']

    doc = SimpleDocTemplate(
        path,
        pagesize=landscape(A3),
        leftMargin=15*mm, rightMargin=15*mm,
        topMargin=15*mm, bottomMargin=15*mm
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title2', parent=styles['Title'], fontSize=18,
                                  textColor=colors.HexColor(PRIMARY))
    subtitle_style = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=10,
                                     textColor=colors.HexColor('#666666'))
    section_style = ParagraphStyle('Section', parent=styles['Heading2'], fontSize=13,
                                    textColor=colors.HexColor(PRIMARY),
                                    spaceAfter=4*mm)

    elements = []
    elements.append(Paragraph('Pagema Ltd &mdash; P&amp;L (No Shareholder) by Branch', title_style))
    elements.append(Paragraph(f'December 2025 &ndash; February 2026 | Generated {get_today()}', subtitle_style))
    elements.append(Spacer(1, 8*mm))

    for month in MONTHS:
        labels, columns = month_table_data(all_data, unassigned, month, section_map)
        elements.append(Paragraph(month, section_style))
        elements.append(build_pdf_table(labels, columns, col_headers))
        elements.append(Spacer(1, 6*mm))

    elements.append(Paragraph('3-Month Summary (Dec 2025 &ndash; Feb 2026)', section_style))
    summary_labels, summary_columns = build_summary(all_data, unassigned, section_map)
    elements.append(build_pdf_table(summary_labels, summary_columns, col_headers))

    doc.build(elements)
    print(f'PDF: {path}')


def build_pdf_table(labels, columns, col_headers):
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import Table, TableStyle

    # Build header
    header = ['Line Item']
    for col in col_headers:
        header.append(col)

    # Build data rows (GBP only for PDF to fit page)
    data = [header]
    row_styles = []  # (row_index, is_total, is_key)

    for i, label in enumerate(labels):
        is_tot = label.startswith('__TOTAL__')
        is_k = label.startswith('__KEY__')
        disp = display_label(label)
        row = [disp]
        for col in col_headers:
            val = columns.get(col, {}).get(label, 0.0)
            ti = get_total_income(columns.get(col, {}))
            pct = fmt_pct(val, ti)
            gbp = fmt_gbp(val)
            row.append(f'{gbp}\n{pct}')
        data.append(row)
        row_styles.append((i + 1, is_tot, is_k))

    col_widths = [55*mm] + [30*mm] * len(col_headers)

    table = Table(data, colWidths=col_widths, repeatRows=1)

    style_cmds = [
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(PRIMARY)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]

    for row_idx, is_tot, is_k in row_styles:
        if is_tot:
            style_cmds.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#E8E8E8')))
            style_cmds.append(('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold'))
        elif is_k:
            style_cmds.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#DCE4ED')))
            style_cmds.append(('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold'))

        # Alternate row colors for non-special rows
        if not is_tot and not is_k and row_idx % 2 == 0:
            style_cmds.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#F9F9F9')))

    # Color negative values red
    for row_idx in range(1, len(data)):
        for col_idx in range(1, len(col_headers) + 1):
            label = labels[row_idx - 1]
            col_name = col_headers[col_idx - 1]
            val = columns.get(col_name, {}).get(label, 0.0)
            if val < 0:
                style_cmds.append(('TEXTCOLOR', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor(RED)))

    table.setStyle(TableStyle(style_cmds))
    return table


# ═══════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════

def main():
    with open(RAW_FILE) as f:
        raw = json.load(f)

    all_data, section_map = build_all_data(raw)
    unassigned = compute_unassigned(all_data)

    generate_html(all_data, unassigned, section_map)
    generate_md(all_data, unassigned, section_map)
    generate_xlsx(all_data, unassigned, section_map)
    generate_pdf(all_data, unassigned, section_map)

    print('\nAll reports generated successfully.')


if __name__ == '__main__':
    main()
