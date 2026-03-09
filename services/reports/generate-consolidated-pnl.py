#!/usr/bin/env python3
"""
IAGTM Consolidated P&L Report Generator
Merges GRL (Paddington) and Pagema (5 dark kitchens) data into a single report.
Outputs: HTML, PDF, Excel, Markdown
"""

import json
import os
from collections import OrderedDict
from datetime import datetime

# ─── Configuration ───────────────────────────────────────────────────────────

BRAND = {
    'primary': '#1B365D',    # dark navy
    'secondary': '#D4A84B',  # gold
    'green': '#2E7D32',      # positive
    'red': '#C62828',        # negative
    'light_bg': '#F5F5F5',
    'white': '#FFFFFF',
    'border': '#D0D0D0',
}

MONTHS_ORDER = [
    ('2025-12', 'December 2025'),
    ('2026-01', 'January 2026'),
    ('2026-02', 'February 2026'),
]

PAGEMA_BRANCHES = ['Brent', 'Chiswick', 'Peckham', 'Shoreditch', 'Wandsworth']
ALL_COLUMNS = ['Paddington'] + PAGEMA_BRANCHES + ['Unassigned', 'TOTAL']

SECTION_MAP = {
    'Income': 'Income',
    'Less Cost of Sales': 'Less Cost of Sales',
    'Plus Other Income': 'Plus Other Income',
    'Less Operating Expenses': 'Less Operating Expenses',
}

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))


# ─── Data Loading ────────────────────────────────────────────────────────────

def load_grl():
    with open(os.path.join(OUTPUT_DIR, 'grl-pnl-paddington-dec25-feb26.json')) as f:
        return json.load(f)


def load_pagema():
    with open(os.path.join(OUTPUT_DIR, 'pagema-pnl-raw.json')) as f:
        return json.load(f)


# ─── Parsing ─────────────────────────────────────────────────────────────────

def parse_grl_month(month_data):
    """Parse GRL month data into {section: {account: amount}}"""
    result = {}
    for section in month_data['sections']:
        title = section['title']
        if title in SECTION_MAP:
            section_key = SECTION_MAP[title]
            accounts = {}
            for row in section['rows']:
                accounts[row['account']] = row['amount']
            result[section_key] = accounts
    return result


def parse_xero_report(report_data):
    """Parse a Xero report response into {section: {account: amount}}"""
    result = {}
    rows = report_data['Reports'][0]['Rows']
    for row in rows:
        if row['RowType'] == 'Section':
            title = row.get('Title', '')
            if title in SECTION_MAP:
                section_key = SECTION_MAP[title]
                accounts = {}
                for r in row.get('Rows', []):
                    if r['RowType'] == 'Row':
                        name = r['Cells'][0]['Value']
                        val_str = r['Cells'][1]['Value']
                        try:
                            accounts[name] = float(val_str)
                        except (ValueError, TypeError):
                            accounts[name] = 0.0
                result[section_key] = accounts
    return result


# ─── Merging ─────────────────────────────────────────────────────────────────

def build_consolidated():
    """Build the full consolidated data structure."""
    grl = load_grl()
    pagema = load_pagema()

    # Map month keys between GRL and Pagema
    month_map = {
        '2025-12': 'December 2025',
        '2026-01': 'January 2026',
        '2026-02': 'February 2026',
    }

    sections_order = ['Income', 'Less Cost of Sales', 'Plus Other Income', 'Less Operating Expenses']

    # Collect all unique accounts per section across all months and all sources
    all_accounts = {s: set() for s in sections_order}

    all_parsed = {}  # {month_key: {column: {section: {account: amount}}}}

    for month_key, month_label in MONTHS_ORDER:
        all_parsed[month_key] = {}

        # GRL / Paddington
        if month_key in grl['months']:
            parsed = parse_grl_month(grl['months'][month_key])
            all_parsed[month_key]['Paddington'] = parsed
            for sec, accts in parsed.items():
                all_accounts[sec].update(accts.keys())

        # Pagema branches + Total
        pagema_label = month_map[month_key]
        if pagema_label in pagema:
            for branch_key in ['Total'] + PAGEMA_BRANCHES:
                if branch_key in pagema[pagema_label]:
                    parsed = parse_xero_report(pagema[pagema_label][branch_key])
                    col = f'Pagema_{branch_key}'
                    all_parsed[month_key][col] = parsed
                    for sec, accts in parsed.items():
                        all_accounts[sec].update(accts.keys())

    # Sort accounts alphabetically within each section
    sorted_accounts = {s: sorted(all_accounts[s]) for s in sections_order}

    # Build monthly tables
    monthly_tables = {}
    for month_key, month_label in MONTHS_ORDER:
        table = build_month_table(all_parsed[month_key], sections_order, sorted_accounts)
        monthly_tables[month_key] = table

    # Build 3-month summary
    summary_table = build_summary_table(monthly_tables, sections_order, sorted_accounts)

    # Build entity breakdown
    entity_breakdown = build_entity_breakdown(monthly_tables)

    return {
        'monthly': monthly_tables,
        'summary': summary_table,
        'entity_breakdown': entity_breakdown,
        'sections_order': sections_order,
        'sorted_accounts': sorted_accounts,
    }


def build_month_table(month_parsed, sections_order, sorted_accounts):
    """Build a table for one month.

    Returns: {
        'sections': {section_name: {'rows': [{account, values: {col: amount}}], 'total': {col: amount}}},
        'gross_profit': {col: amount},
        'net_profit': {col: amount},
        'total_income': {col: amount},  # for % calculations
    }
    """
    table = {'sections': {}}

    for section in sections_order:
        rows = []
        section_totals = {col: 0.0 for col in ALL_COLUMNS}

        for account in sorted_accounts[section]:
            values = {}

            # Paddington (GRL)
            pad_val = month_parsed.get('Paddington', {}).get(section, {}).get(account, 0.0)
            values['Paddington'] = pad_val

            # Pagema branches
            branch_sum = 0.0
            for branch in PAGEMA_BRANCHES:
                bval = month_parsed.get(f'Pagema_{branch}', {}).get(section, {}).get(account, 0.0)
                values[branch] = bval
                branch_sum += bval

            # Pagema Total
            pagema_total_val = month_parsed.get('Pagema_Total', {}).get(section, {}).get(account, 0.0)

            # Unassigned = Pagema Total - sum of branches
            values['Unassigned'] = pagema_total_val - branch_sum

            # TOTAL = Paddington + Pagema Total
            values['TOTAL'] = pad_val + pagema_total_val

            # Only add row if any value is non-zero
            if any(abs(v) > 0.005 for v in values.values()):
                rows.append({'account': account, 'values': values})
                for col in ALL_COLUMNS:
                    section_totals[col] += values[col]

        table['sections'][section] = {
            'rows': rows,
            'total': section_totals,
        }

    # Calculate derived rows
    income_total = table['sections']['Income']['total']
    cogs_total = table['sections']['Less Cost of Sales']['total']
    other_income_total = table['sections']['Plus Other Income']['total']
    opex_total = table['sections']['Less Operating Expenses']['total']

    gross_profit = {col: income_total[col] - cogs_total[col] for col in ALL_COLUMNS}
    net_profit = {col: gross_profit[col] + other_income_total[col] - opex_total[col] for col in ALL_COLUMNS}

    table['gross_profit'] = gross_profit
    table['net_profit'] = net_profit
    table['total_income'] = dict(income_total)

    return table


def build_summary_table(monthly_tables, sections_order, sorted_accounts):
    """Sum all months into a 3-month summary."""
    summary = {
        'sections': {},
        'gross_profit': {col: 0.0 for col in ALL_COLUMNS},
        'net_profit': {col: 0.0 for col in ALL_COLUMNS},
        'total_income': {col: 0.0 for col in ALL_COLUMNS},
    }

    for section in sections_order:
        # Merge rows across months
        account_values = {}
        for month_key, _ in MONTHS_ORDER:
            mt = monthly_tables[month_key]
            for row in mt['sections'][section]['rows']:
                acct = row['account']
                if acct not in account_values:
                    account_values[acct] = {col: 0.0 for col in ALL_COLUMNS}
                for col in ALL_COLUMNS:
                    account_values[acct][col] += row['values'][col]

        rows = []
        section_totals = {col: 0.0 for col in ALL_COLUMNS}
        for acct in sorted_accounts[section]:
            if acct in account_values:
                rows.append({'account': acct, 'values': account_values[acct]})
                for col in ALL_COLUMNS:
                    section_totals[col] += account_values[acct][col]

        summary['sections'][section] = {'rows': rows, 'total': section_totals}

    for col in ALL_COLUMNS:
        for month_key, _ in MONTHS_ORDER:
            summary['gross_profit'][col] += monthly_tables[month_key]['gross_profit'][col]
            summary['net_profit'][col] += monthly_tables[month_key]['net_profit'][col]
            summary['total_income'][col] += monthly_tables[month_key]['total_income'][col]

    return summary


def build_entity_breakdown(monthly_tables):
    """Build entity-level summary: GRL | Pagema | IAGTM Total"""
    breakdown = {}
    for period_key, period_label in MONTHS_ORDER + [('summary', '3-Month Total')]:
        if period_key == 'summary':
            # We'll compute this after
            continue

        mt = monthly_tables[period_key]
        grl_income = mt['total_income']['Paddington']
        pagema_income = mt['total_income']['TOTAL'] - mt['total_income']['Paddington']
        total_income = mt['total_income']['TOTAL']

        grl_cogs = mt['sections']['Less Cost of Sales']['total']['Paddington']
        pagema_cogs = mt['sections']['Less Cost of Sales']['total']['TOTAL'] - mt['sections']['Less Cost of Sales']['total']['Paddington']
        total_cogs = mt['sections']['Less Cost of Sales']['total']['TOTAL']

        grl_gp = mt['gross_profit']['Paddington']
        pagema_gp = mt['gross_profit']['TOTAL'] - mt['gross_profit']['Paddington']
        total_gp = mt['gross_profit']['TOTAL']

        grl_opex = mt['sections']['Less Operating Expenses']['total']['Paddington']
        pagema_opex = mt['sections']['Less Operating Expenses']['total']['TOTAL'] - mt['sections']['Less Operating Expenses']['total']['Paddington']
        total_opex = mt['sections']['Less Operating Expenses']['total']['TOTAL']

        grl_np = mt['net_profit']['Paddington']
        pagema_np = mt['net_profit']['TOTAL'] - mt['net_profit']['Paddington']
        total_np = mt['net_profit']['TOTAL']

        breakdown[period_key] = {
            'label': period_label,
            'rows': [
                ('Total Income', grl_income, pagema_income, total_income),
                ('Total COGS', grl_cogs, pagema_cogs, total_cogs),
                ('Gross Profit', grl_gp, pagema_gp, total_gp),
                ('Operating Expenses', grl_opex, pagema_opex, total_opex),
                ('Net Profit', grl_np, pagema_np, total_np),
            ]
        }

    # 3-month total
    totals = {metric: [0.0, 0.0, 0.0] for metric in ['Total Income', 'Total COGS', 'Gross Profit', 'Operating Expenses', 'Net Profit']}
    for mk, _ in MONTHS_ORDER:
        for row in breakdown[mk]['rows']:
            totals[row[0]][0] += row[1]
            totals[row[0]][1] += row[2]
            totals[row[0]][2] += row[3]

    breakdown['summary'] = {
        'label': '3-Month Total (Dec 2025 - Feb 2026)',
        'rows': [(k, v[0], v[1], v[2]) for k, v in totals.items()]
    }

    return breakdown


# ─── Formatting Helpers ──────────────────────────────────────────────────────

def fmt_currency(val):
    """Format as GBP with commas, negatives in parentheses."""
    if val is None:
        return '-'
    if abs(val) < 0.005:
        return '-'
    if val < 0:
        return f'({abs(val):,.2f})'
    return f'{val:,.2f}'


def fmt_pct(val, total_income):
    """Format as percentage of total income."""
    if total_income is None or abs(total_income) < 0.01:
        return '-'
    if val is None or abs(val) < 0.005:
        return '-'
    pct = (val / total_income) * 100
    if pct < 0:
        return f'({abs(pct):.1f}%)'
    return f'{pct:.1f}%'


# ─── HTML Generation ─────────────────────────────────────────────────────────

def generate_html(data):
    section_labels = {
        'Income': ('Income', 'Total Income'),
        'Less Cost of Sales': ('Cost of Sales', 'Total Cost of Sales'),
        'Plus Other Income': ('Other Income', 'Total Other Income'),
        'Less Operating Expenses': ('Operating Expenses', 'Total Operating Expenses'),
    }

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>IAGTM - Consolidated P&L (All Locations) - Dec 2025 to Feb 2026</title>
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{ font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif; background: {BRAND['light_bg']}; color: #333; padding: 20px; }}

  .report-header {{
    background: {BRAND['primary']};
    color: white;
    padding: 24px 32px;
    border-radius: 8px 8px 0 0;
    margin-bottom: 0;
  }}
  .report-header h1 {{ font-size: 22px; font-weight: 700; margin-bottom: 4px; }}
  .report-header .subtitle {{ font-size: 13px; opacity: 0.85; }}

  .report-container {{ max-width: 1600px; margin: 0 auto; }}

  .month-title {{
    background: {BRAND['secondary']};
    color: {BRAND['primary']};
    padding: 12px 20px;
    font-size: 16px;
    font-weight: 700;
    margin-top: 32px;
    border-radius: 6px 6px 0 0;
  }}

  .table-wrap {{ overflow-x: auto; margin-bottom: 24px; border: 1px solid {BRAND['border']}; border-radius: 0 0 6px 6px; }}

  table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 11px;
    min-width: 1200px;
  }}

  th, td {{ padding: 4px 8px; text-align: right; white-space: nowrap; border-bottom: 1px solid #E8E8E8; }}
  th {{ background: {BRAND['primary']}; color: white; font-weight: 600; font-size: 10px; text-transform: uppercase; letter-spacing: 0.5px; position: sticky; top: 0; z-index: 2; }}
  th:first-child {{ text-align: left; min-width: 240px; position: sticky; left: 0; z-index: 3; }}
  td:first-child {{ text-align: left; font-weight: 400; position: sticky; left: 0; background: inherit; z-index: 1; min-width: 240px; }}

  /* Visual separator between Paddington and dark kitchens */
  th.sep, td.sep {{ border-left: 3px solid {BRAND['secondary']}; }}

  /* Column pairs: amount + pct */
  .pct {{ color: #888; font-size: 10px; }}

  tr:nth-child(even) td {{ background-color: #FAFAFA; }}
  tr:nth-child(odd) td {{ background-color: white; }}
  tr:nth-child(even) td:first-child {{ background-color: #FAFAFA; }}
  tr:nth-child(odd) td:first-child {{ background-color: white; }}

  tr.section-header td {{
    background: {BRAND['primary']}10 !important;
    font-weight: 700;
    font-size: 11px;
    color: {BRAND['primary']};
    border-top: 2px solid {BRAND['primary']};
    padding-top: 8px;
  }}

  tr.section-total td {{
    font-weight: 700;
    border-top: 1px solid #999;
    border-bottom: 2px solid {BRAND['primary']};
    background: #F0F0F0 !important;
  }}
  tr.section-total td:first-child {{ background: #F0F0F0 !important; }}

  tr.grand-total td {{
    font-weight: 700;
    font-size: 12px;
    border-top: 2px solid {BRAND['primary']};
    border-bottom: 2px solid {BRAND['primary']};
    background: {BRAND['secondary']}22 !important;
    color: {BRAND['primary']};
  }}
  tr.grand-total td:first-child {{ background: {BRAND['secondary']}22 !important; }}

  .negative {{ color: {BRAND['red']}; }}
  .positive {{ color: {BRAND['green']}; }}

  .report-footer {{
    margin-top: 16px;
    padding: 16px 20px;
    background: white;
    border: 1px solid {BRAND['border']};
    border-radius: 6px;
    font-size: 11px;
    color: #666;
    line-height: 1.6;
  }}
  .report-footer strong {{ color: {BRAND['primary']}; }}

  .entity-table {{ margin-top: 24px; }}
  .entity-table table {{ min-width: 600px; font-size: 12px; }}
  .entity-table th {{ font-size: 11px; }}
  .entity-period {{ font-weight: 600; color: {BRAND['primary']}; padding: 8px 0 4px; font-size: 13px; }}
</style>
</head>
<body>
<div class="report-container">
  <div class="report-header">
    <h1>IAGTM — Consolidated P&L (All Locations)</h1>
    <div class="subtitle">December 2025 — February 2026 &nbsp;|&nbsp; Generated {datetime.now().strftime('%d %B %Y %H:%M')}</div>
  </div>
"""

    def val_class(val):
        if val is None or abs(val) < 0.005:
            return ''
        return 'negative' if val < 0 else ''

    def render_table(table_data, title=None):
        """Render a full P&L table."""
        out = ''
        if title:
            out += f'<div class="month-title">{title}</div>\n'
        out += '<div class="table-wrap"><table>\n'

        # Header row
        out += '<tr>'
        out += '<th>Account</th>'
        out += '<th colspan="2">Paddington</th>'
        out += f'<th colspan="2" class="sep">Brent</th>'
        out += '<th colspan="2">Chiswick</th>'
        out += '<th colspan="2">Peckham</th>'
        out += '<th colspan="2">Shoreditch</th>'
        out += '<th colspan="2">Wandsworth</th>'
        out += '<th colspan="2">Unassigned</th>'
        out += '<th colspan="2">TOTAL</th>'
        out += '</tr>\n'

        # Sub-header
        out += '<tr>'
        out += '<th></th>'
        out += '<th>&pound;</th><th>%</th>'
        out += f'<th class="sep">&pound;</th><th>%</th>'
        for _ in range(6):
            out += '<th>&pound;</th><th>%</th>'
        out += '</tr>\n'

        ti = table_data['total_income']

        for section_key in data['sections_order']:
            label, total_label = section_labels[section_key]
            sec = table_data['sections'][section_key]

            # Section header row
            out += f'<tr class="section-header"><td>{label}</td>' + '<td></td>' * 16 + '</tr>\n'

            # Data rows
            for row in sec['rows']:
                acct = row['account']
                vals = row['values']
                out += '<tr>'
                out += f'<td>&nbsp;&nbsp;{acct}</td>'
                for i, col in enumerate(ALL_COLUMNS):
                    v = vals[col]
                    cls = val_class(v)
                    sep = ' class="sep"' if i == 1 else ''  # Brent is index 1
                    if i == 1:
                        out += f'<td class="sep {cls}">{fmt_currency(v)}</td>'
                    else:
                        out += f'<td class="{cls}">{fmt_currency(v)}</td>'
                    out += f'<td class="pct">{fmt_pct(v, ti[col])}</td>'
                out += '</tr>\n'

            # Section total
            out += f'<tr class="section-total"><td>{total_label}</td>'
            for i, col in enumerate(ALL_COLUMNS):
                v = sec['total'][col]
                cls = val_class(v)
                if i == 1:
                    out += f'<td class="sep {cls}">{fmt_currency(v)}</td>'
                else:
                    out += f'<td class="{cls}">{fmt_currency(v)}</td>'
                out += f'<td class="pct">{fmt_pct(v, ti[col])}</td>'
            out += '</tr>\n'

            # After COGS: Gross Profit
            if section_key == 'Less Cost of Sales':
                out += '<tr class="grand-total"><td>Gross Profit</td>'
                for i, col in enumerate(ALL_COLUMNS):
                    v = table_data['gross_profit'][col]
                    cls = val_class(v)
                    if i == 1:
                        out += f'<td class="sep {cls}">{fmt_currency(v)}</td>'
                    else:
                        out += f'<td class="{cls}">{fmt_currency(v)}</td>'
                    out += f'<td class="pct">{fmt_pct(v, ti[col])}</td>'
                out += '</tr>\n'

        # Net Profit
        out += '<tr class="grand-total"><td>Net Profit</td>'
        for i, col in enumerate(ALL_COLUMNS):
            v = table_data['net_profit'][col]
            cls = val_class(v)
            if i == 1:
                out += f'<td class="sep {cls}">{fmt_currency(v)}</td>'
            else:
                out += f'<td class="{cls}">{fmt_currency(v)}</td>'
            out += f'<td class="pct">{fmt_pct(v, ti[col])}</td>'
        out += '</tr>\n'

        out += '</table></div>\n'
        return out

    # Monthly tables
    for month_key, month_label in MONTHS_ORDER:
        html += render_table(data['monthly'][month_key], month_label)

    # 3-Month Summary
    html += render_table(data['summary'], '3-Month Consolidated Summary (Dec 2025 - Feb 2026)')

    # Entity Breakdown
    html += '<div class="entity-table">\n'
    html += '<div class="month-title">Entity Breakdown Summary</div>\n'
    html += '<div class="table-wrap"><table>\n'
    html += '<tr><th style="text-align:left">Period / Metric</th><th>GRL (Paddington)</th><th>Pagema (Dark Kitchens)</th><th>IAGTM Total</th></tr>\n'

    for period_key in [mk for mk, _ in MONTHS_ORDER] + ['summary']:
        bd = data['entity_breakdown'][period_key]
        html += f'<tr class="section-header"><td colspan="4">{bd["label"]}</td></tr>\n'
        for metric, grl_val, pagema_val, total_val in bd['rows']:
            cls_g = val_class(grl_val)
            cls_p = val_class(pagema_val)
            cls_t = val_class(total_val)
            bold = ' style="font-weight:700"' if metric in ('Gross Profit', 'Net Profit') else ''
            html += f'<tr{bold}>'
            html += f'<td>&nbsp;&nbsp;{metric}</td>'
            html += f'<td class="{cls_g}">&pound;{fmt_currency(grl_val)}</td>'
            html += f'<td class="{cls_p}">&pound;{fmt_currency(pagema_val)}</td>'
            html += f'<td class="{cls_t}">&pound;{fmt_currency(total_val)}</td>'
            html += '</tr>\n'

    html += '</table></div></div>\n'

    # Footer
    html += f"""
  <div class="report-footer">
    <strong>GRL</strong> = Greek Restaurant Ltd (Paddington) &nbsp;|&nbsp;
    <strong>Pagema Ltd</strong> = Dark Kitchens (Brent, Chiswick, Peckham, Shoreditch, Wandsworth)<br>
    <strong>Unassigned</strong> = Pagema Total minus sum of individual branches (rounding / unallocated items)<br>
    <strong>%</strong> = Percentage of that column's Total Income<br><br>
    <em>Note: February 2026 Pagema expenses appear incomplete — likely pending Xero reconciliation.</em>
  </div>
</div>
</body>
</html>
"""

    path = os.path.join(OUTPUT_DIR, 'iagtm-consolidated-pnl-dec25-feb26.html')
    with open(path, 'w') as f:
        f.write(html)
    print(f'  HTML: {path}')
    return path


# ─── Markdown Generation ─────────────────────────────────────────────────────

def generate_markdown(data):
    section_labels = {
        'Income': ('Income', 'Total Income'),
        'Less Cost of Sales': ('Cost of Sales', 'Total Cost of Sales'),
        'Plus Other Income': ('Other Income', 'Total Other Income'),
        'Less Operating Expenses': ('Operating Expenses', 'Total Operating Expenses'),
    }

    lines = []
    lines.append('# IAGTM - Consolidated P&L (All Locations)')
    lines.append(f'**December 2025 - February 2026** | Generated {datetime.now().strftime("%d %B %Y %H:%M")}')
    lines.append('')

    def render_md_table(table_data, title):
        out = []
        out.append(f'## {title}')
        out.append('')
        # Header (just amounts, no % for markdown — too wide)
        cols = ALL_COLUMNS
        out.append('| Account | ' + ' | '.join(cols) + ' |')
        out.append('|:--------|' + '|'.join(['---------:'] * len(cols)) + '|')

        ti = table_data['total_income']

        for section_key in data['sections_order']:
            label, total_label = section_labels[section_key]
            sec = table_data['sections'][section_key]

            out.append(f'| **{label}** |' + ' |' * len(cols) + '')

            for row in sec['rows']:
                vals = row['values']
                cells = [fmt_currency(vals[c]) for c in cols]
                out.append(f'| {row["account"]} | ' + ' | '.join(cells) + ' |')

            # Total
            tot_cells = [fmt_currency(sec['total'][c]) for c in cols]
            out.append(f'| **{total_label}** | ' + ' | '.join(f'**{c}**' for c in tot_cells) + ' |')

            if section_key == 'Less Cost of Sales':
                gp_cells = [fmt_currency(table_data['gross_profit'][c]) for c in cols]
                out.append(f'| **Gross Profit** | ' + ' | '.join(f'**{c}**' for c in gp_cells) + ' |')

        np_cells = [fmt_currency(table_data['net_profit'][c]) for c in cols]
        out.append(f'| **Net Profit** | ' + ' | '.join(f'**{c}**' for c in np_cells) + ' |')
        out.append('')
        return out

    for month_key, month_label in MONTHS_ORDER:
        lines.extend(render_md_table(data['monthly'][month_key], month_label))

    lines.extend(render_md_table(data['summary'], '3-Month Consolidated Summary'))

    # Entity breakdown
    lines.append('## Entity Breakdown Summary')
    lines.append('')
    lines.append('| Period / Metric | GRL (Paddington) | Pagema (Dark Kitchens) | IAGTM Total |')
    lines.append('|:----------------|------------------:|----------------------:|-----------:|')

    for period_key in [mk for mk, _ in MONTHS_ORDER] + ['summary']:
        bd = data['entity_breakdown'][period_key]
        lines.append(f'| **{bd["label"]}** | | | |')
        for metric, grl_val, pagema_val, total_val in bd['rows']:
            lines.append(f'| {metric} | {fmt_currency(grl_val)} | {fmt_currency(pagema_val)} | {fmt_currency(total_val)} |')

    lines.append('')
    lines.append('---')
    lines.append('*Note: February 2026 Pagema expenses appear incomplete - likely pending Xero reconciliation.*')
    lines.append('')
    lines.append('GRL = Greek Restaurant Ltd (Paddington) | Pagema Ltd = Dark Kitchens (Brent, Chiswick, Peckham, Shoreditch, Wandsworth)')

    path = os.path.join(OUTPUT_DIR, 'iagtm-consolidated-pnl-dec25-feb26.md')
    with open(path, 'w') as f:
        f.write('\n'.join(lines))
    print(f'  Markdown: {path}')
    return path


# ─── Excel Generation ────────────────────────────────────────────────────────

def generate_excel(data):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    section_labels = {
        'Income': ('Income', 'Total Income'),
        'Less Cost of Sales': ('Cost of Sales', 'Total Cost of Sales'),
        'Plus Other Income': ('Other Income', 'Total Other Income'),
        'Less Operating Expenses': ('Operating Expenses', 'Total Operating Expenses'),
    }

    # Styles
    header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1B365D', end_color='1B365D', fill_type='solid')
    gold_fill = PatternFill(start_color='D4A84B', end_color='D4A84B', fill_type='solid')
    section_font = Font(name='Calibri', size=10, bold=True, color='1B365D')
    section_fill = PatternFill(start_color='E8EDF3', end_color='E8EDF3', fill_type='solid')
    total_font = Font(name='Calibri', size=10, bold=True)
    total_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    grand_font = Font(name='Calibri', size=11, bold=True, color='1B365D')
    grand_fill = PatternFill(start_color='FDF6E3', end_color='FDF6E3', fill_type='solid')
    normal_font = Font(name='Calibri', size=10)
    red_font = Font(name='Calibri', size=10, color='C62828')
    pct_font = Font(name='Calibri', size=9, color='888888')
    thin_border = Border(bottom=Side(style='thin', color='D0D0D0'))
    thick_border = Border(bottom=Side(style='medium', color='1B365D'))

    gbp_fmt = '#,##0.00;(#,##0.00);"-"'
    pct_fmt = '0.0%;(0.0%);"-"'

    def write_sheet(ws, table_data, title):
        ws.sheet_properties.tabColor = '1B365D'
        ws.freeze_panes = 'B4'

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=17)
        c = ws.cell(row=1, column=1, value=f'IAGTM Consolidated P&L - {title}')
        c.font = Font(name='Calibri', size=14, bold=True, color='1B365D')

        # Column headers (row 2: location names, row 3: GBP/%)
        col_map = {}  # column_name -> (gbp_col_idx, pct_col_idx)
        col_idx = 2
        for i, col_name in enumerate(ALL_COLUMNS):
            col_map[col_name] = (col_idx, col_idx + 1)

            ws.merge_cells(start_row=2, start_column=col_idx, end_row=2, end_column=col_idx + 1)
            c = ws.cell(row=2, column=col_idx, value=col_name)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal='center')
            ws.cell(row=2, column=col_idx + 1).fill = header_fill

            c1 = ws.cell(row=3, column=col_idx, value='GBP')
            c1.font = header_font
            c1.fill = header_fill
            c1.alignment = Alignment(horizontal='right')
            c2 = ws.cell(row=3, column=col_idx + 1, value='%')
            c2.font = header_font
            c2.fill = header_fill
            c2.alignment = Alignment(horizontal='right')

            col_idx += 2

        # Account header
        for r in [2, 3]:
            c = ws.cell(row=r, column=1)
            c.fill = header_fill
            c.font = header_font
        ws.cell(row=2, column=1, value='Account').font = header_font
        ws.cell(row=3, column=1, value='').font = header_font

        ti = table_data['total_income']

        row_num = 4

        for section_key in data['sections_order']:
            label, total_label = section_labels[section_key]
            sec = table_data['sections'][section_key]

            # Section header
            c = ws.cell(row=row_num, column=1, value=label)
            c.font = section_font
            c.fill = section_fill
            for ci in range(2, col_idx):
                ws.cell(row=row_num, column=ci).fill = section_fill
            row_num += 1

            # Data rows
            for row in sec['rows']:
                ws.cell(row=row_num, column=1, value=f'  {row["account"]}').font = normal_font
                for col_name in ALL_COLUMNS:
                    gbp_col, pct_col = col_map[col_name]
                    v = row['values'][col_name]
                    cell = ws.cell(row=row_num, column=gbp_col, value=v if abs(v) > 0.005 else None)
                    cell.number_format = gbp_fmt
                    cell.font = red_font if v < -0.005 else normal_font
                    cell.border = thin_border

                    pct_val = (v / ti[col_name]) if ti[col_name] and abs(ti[col_name]) > 0.01 and abs(v) > 0.005 else None
                    pc = ws.cell(row=row_num, column=pct_col, value=pct_val)
                    pc.number_format = pct_fmt
                    pc.font = pct_font
                    pc.border = thin_border
                row_num += 1

            # Section total
            ws.cell(row=row_num, column=1, value=total_label).font = total_font
            for ci in range(1, col_idx):
                ws.cell(row=row_num, column=ci).fill = total_fill
                ws.cell(row=row_num, column=ci).border = thick_border
            for col_name in ALL_COLUMNS:
                gbp_col, pct_col = col_map[col_name]
                v = sec['total'][col_name]
                cell = ws.cell(row=row_num, column=gbp_col, value=v if abs(v) > 0.005 else None)
                cell.number_format = gbp_fmt
                cell.font = total_font
                pct_val = (v / ti[col_name]) if ti[col_name] and abs(ti[col_name]) > 0.01 and abs(v) > 0.005 else None
                pc = ws.cell(row=row_num, column=pct_col, value=pct_val)
                pc.number_format = pct_fmt
                pc.font = Font(name='Calibri', size=9, bold=True, color='888888')
            row_num += 1

            # Gross Profit after COGS
            if section_key == 'Less Cost of Sales':
                ws.cell(row=row_num, column=1, value='Gross Profit').font = grand_font
                for ci in range(1, col_idx):
                    ws.cell(row=row_num, column=ci).fill = grand_fill
                    ws.cell(row=row_num, column=ci).border = thick_border
                for col_name in ALL_COLUMNS:
                    gbp_col, pct_col = col_map[col_name]
                    v = table_data['gross_profit'][col_name]
                    cell = ws.cell(row=row_num, column=gbp_col, value=v)
                    cell.number_format = gbp_fmt
                    cell.font = grand_font
                    pct_val = (v / ti[col_name]) if ti[col_name] and abs(ti[col_name]) > 0.01 else None
                    pc = ws.cell(row=row_num, column=pct_col, value=pct_val)
                    pc.number_format = pct_fmt
                    pc.font = Font(name='Calibri', size=9, bold=True, color='1B365D')
                row_num += 1

        # Net Profit
        ws.cell(row=row_num, column=1, value='Net Profit').font = grand_font
        for ci in range(1, col_idx):
            ws.cell(row=row_num, column=ci).fill = grand_fill
            ws.cell(row=row_num, column=ci).border = thick_border
        for col_name in ALL_COLUMNS:
            gbp_col, pct_col = col_map[col_name]
            v = table_data['net_profit'][col_name]
            cell = ws.cell(row=row_num, column=gbp_col, value=v)
            cell.number_format = gbp_fmt
            cell.font = grand_font
            pct_val = (v / ti[col_name]) if ti[col_name] and abs(ti[col_name]) > 0.01 else None
            pc = ws.cell(row=row_num, column=pct_col, value=pct_val)
            pc.number_format = pct_fmt
            pc.font = Font(name='Calibri', size=9, bold=True, color='1B365D')
        row_num += 1

        # Footnote
        row_num += 1
        ws.cell(row=row_num, column=1, value='Note: February 2026 Pagema expenses appear incomplete - likely pending Xero reconciliation.').font = Font(name='Calibri', size=9, italic=True, color='888888')

        # Column widths
        ws.column_dimensions['A'].width = 35
        for ci in range(2, col_idx):
            ws.column_dimensions[get_column_letter(ci)].width = 12

    # Sheet 1-3: Monthly
    sheets_created = False
    for month_key, month_label in MONTHS_ORDER:
        if not sheets_created:
            ws = wb.active
            ws.title = month_label
            sheets_created = True
        else:
            ws = wb.create_sheet(title=month_label)
        write_sheet(ws, data['monthly'][month_key], month_label)

    # Sheet 4: 3-Month Summary
    ws = wb.create_sheet(title='3-Month Summary')
    write_sheet(ws, data['summary'], '3-Month Summary (Dec 2025 - Feb 2026)')

    # Sheet 5: Entity Breakdown
    ws = wb.create_sheet(title='Entity Breakdown')
    ws.sheet_properties.tabColor = 'D4A84B'
    ws.freeze_panes = 'B3'

    ws.merge_cells('A1:D1')
    c = ws.cell(row=1, column=1, value='IAGTM Entity Breakdown Summary')
    c.font = Font(name='Calibri', size=14, bold=True, color='1B365D')

    headers = ['Period / Metric', 'GRL (Paddington)', 'Pagema (Dark Kitchens)', 'IAGTM Total']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='right' if i > 1 else 'left')

    row_num = 3
    for period_key in [mk for mk, _ in MONTHS_ORDER] + ['summary']:
        bd = data['entity_breakdown'][period_key]
        c = ws.cell(row=row_num, column=1, value=bd['label'])
        c.font = section_font
        c.fill = section_fill
        for ci in range(2, 5):
            ws.cell(row=row_num, column=ci).fill = section_fill
        row_num += 1

        for metric, grl_val, pagema_val, total_val in bd['rows']:
            is_key = metric in ('Gross Profit', 'Net Profit')
            fnt = total_font if is_key else normal_font
            fl = grand_fill if is_key else None

            ws.cell(row=row_num, column=1, value=f'  {metric}').font = fnt
            for ci, val in [(2, grl_val), (3, pagema_val), (4, total_val)]:
                cell = ws.cell(row=row_num, column=ci, value=val)
                cell.number_format = gbp_fmt
                cell.font = Font(name='Calibri', size=10, bold=is_key, color='C62828' if val < -0.005 else '1B365D' if is_key else '333333')
                if fl:
                    cell.fill = fl
            row_num += 1

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 18

    path = os.path.join(OUTPUT_DIR, 'iagtm-consolidated-pnl-dec25-feb26.xlsx')
    wb.save(path)
    print(f'  Excel: {path}')
    return path


# ─── PDF Generation ──────────────────────────────────────────────────────────

def generate_pdf(data):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm, cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER

    path = os.path.join(OUTPUT_DIR, 'iagtm-consolidated-pnl-dec25-feb26.pdf')

    page_w, page_h = landscape(A3)
    doc = SimpleDocTemplate(
        path,
        pagesize=landscape(A3),
        leftMargin=12*mm, rightMargin=12*mm,
        topMargin=15*mm, bottomMargin=15*mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('TitleCustom', parent=styles['Title'], fontSize=18, textColor=colors.HexColor('#1B365D'), spaceAfter=4*mm)
    subtitle_style = ParagraphStyle('SubCustom', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#666666'), spaceAfter=6*mm)
    section_title_style = ParagraphStyle('SectionTitle', parent=styles['Heading2'], fontSize=14, textColor=colors.HexColor('#1B365D'), spaceBefore=8*mm, spaceAfter=3*mm)
    footnote_style = ParagraphStyle('Footnote', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#888888'), spaceBefore=4*mm, italic=True)

    section_labels = {
        'Income': ('Income', 'Total Income'),
        'Less Cost of Sales': ('Cost of Sales', 'Total Cost of Sales'),
        'Plus Other Income': ('Other Income', 'Total Other Income'),
        'Less Operating Expenses': ('Operating Expenses', 'Total Operating Expenses'),
    }

    navy = colors.HexColor('#1B365D')
    gold = colors.HexColor('#D4A84B')
    light_blue = colors.HexColor('#E8EDF3')
    light_gold = colors.HexColor('#FDF6E3')
    light_gray = colors.HexColor('#F0F0F0')
    red_c = colors.HexColor('#C62828')

    elements = []
    elements.append(Paragraph('IAGTM — Consolidated P&L (All Locations)', title_style))
    elements.append(Paragraph(f'December 2025 — February 2026 | Generated {datetime.now().strftime("%d %B %Y %H:%M")}', subtitle_style))

    def build_pdf_table(table_data, title):
        """Build a reportlab table for one period."""
        elems = []
        elems.append(Paragraph(title, section_title_style))

        # We'll use just GBP columns (no % — too wide even for A3)
        header = ['Account'] + ALL_COLUMNS
        tdata = [header]

        ti = table_data['total_income']

        row_styles = []  # (row_idx, style_type)
        row_idx = 1

        for section_key in data['sections_order']:
            label, total_label = section_labels[section_key]
            sec = table_data['sections'][section_key]

            tdata.append([label] + [''] * len(ALL_COLUMNS))
            row_styles.append((row_idx, 'section'))
            row_idx += 1

            for row in sec['rows']:
                vals = row['values']
                cells = [f'  {row["account"]}'] + [fmt_currency(vals[c]) for c in ALL_COLUMNS]
                tdata.append(cells)
                row_styles.append((row_idx, 'normal'))
                row_idx += 1

            tot_cells = [total_label] + [fmt_currency(sec['total'][c]) for c in ALL_COLUMNS]
            tdata.append(tot_cells)
            row_styles.append((row_idx, 'total'))
            row_idx += 1

            if section_key == 'Less Cost of Sales':
                gp_cells = ['Gross Profit'] + [fmt_currency(table_data['gross_profit'][c]) for c in ALL_COLUMNS]
                tdata.append(gp_cells)
                row_styles.append((row_idx, 'grand'))
                row_idx += 1

        np_cells = ['Net Profit'] + [fmt_currency(table_data['net_profit'][c]) for c in ALL_COLUMNS]
        tdata.append(np_cells)
        row_styles.append((row_idx, 'grand'))

        # Column widths
        avail = page_w - 24*mm
        acct_w = 120*mm
        data_w = (avail - acct_w) / len(ALL_COLUMNS)
        col_widths = [acct_w] + [data_w] * len(ALL_COLUMNS)

        t = Table(tdata, colWidths=col_widths, repeatRows=1)

        style_cmds = [
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), navy),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D0D0D0')),
            # Separator between Paddington and Brent
            ('LINEAFTER', (1, 0), (1, -1), 1.5, gold),
        ]

        for ri, stype in row_styles:
            if stype == 'section':
                style_cmds.append(('BACKGROUND', (0, ri), (-1, ri), light_blue))
                style_cmds.append(('FONTNAME', (0, ri), (-1, ri), 'Helvetica-Bold'))
                style_cmds.append(('TEXTCOLOR', (0, ri), (-1, ri), navy))
            elif stype == 'total':
                style_cmds.append(('BACKGROUND', (0, ri), (-1, ri), light_gray))
                style_cmds.append(('FONTNAME', (0, ri), (-1, ri), 'Helvetica-Bold'))
                style_cmds.append(('LINEABOVE', (0, ri), (-1, ri), 0.75, navy))
            elif stype == 'grand':
                style_cmds.append(('BACKGROUND', (0, ri), (-1, ri), light_gold))
                style_cmds.append(('FONTNAME', (0, ri), (-1, ri), 'Helvetica-Bold'))
                style_cmds.append(('TEXTCOLOR', (0, ri), (-1, ri), navy))
                style_cmds.append(('LINEABOVE', (0, ri), (-1, ri), 1, navy))
                style_cmds.append(('LINEBELOW', (0, ri), (-1, ri), 1, navy))

        # Color negative values red
        for ri in range(1, len(tdata)):
            for ci in range(1, len(tdata[ri])):
                val = tdata[ri][ci]
                if isinstance(val, str) and val.startswith('('):
                    style_cmds.append(('TEXTCOLOR', (ci, ri), (ci, ri), red_c))

        t.setStyle(TableStyle(style_cmds))
        elems.append(t)
        return elems

    # Monthly tables
    for i, (month_key, month_label) in enumerate(MONTHS_ORDER):
        elements.extend(build_pdf_table(data['monthly'][month_key], month_label))
        if i < len(MONTHS_ORDER) - 1:
            elements.append(PageBreak())

    elements.append(PageBreak())
    elements.extend(build_pdf_table(data['summary'], '3-Month Consolidated Summary'))

    # Entity breakdown
    elements.append(PageBreak())
    elements.append(Paragraph('Entity Breakdown Summary', section_title_style))

    for period_key in [mk for mk, _ in MONTHS_ORDER] + ['summary']:
        bd = data['entity_breakdown'][period_key]
        tdata = [['', 'GRL (Paddington)', 'Pagema (Dark Kitchens)', 'IAGTM Total']]
        for metric, grl_val, pagema_val, total_val in bd['rows']:
            tdata.append([metric, fmt_currency(grl_val), fmt_currency(pagema_val), fmt_currency(total_val)])

        col_widths = [120*mm, 80*mm, 80*mm, 80*mm]
        t = Table(tdata, colWidths=col_widths)
        style_cmds = [
            ('BACKGROUND', (0, 0), (-1, 0), navy),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D0D0D0')),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]
        for ri, row in enumerate(tdata):
            if ri > 0 and row[0] in ('Gross Profit', 'Net Profit'):
                style_cmds.append(('FONTNAME', (0, ri), (-1, ri), 'Helvetica-Bold'))
                style_cmds.append(('BACKGROUND', (0, ri), (-1, ri), light_gold))
            # Color negatives
            for ci in range(1, 4):
                if isinstance(row[ci], str) and row[ci].startswith('('):
                    style_cmds.append(('TEXTCOLOR', (ci, ri), (ci, ri), red_c))

        t.setStyle(TableStyle(style_cmds))

        elements.append(Spacer(1, 4*mm))
        elements.append(Paragraph(f'<b>{bd["label"]}</b>', ParagraphStyle('P', fontSize=10, textColor=navy)))
        elements.append(Spacer(1, 2*mm))
        elements.append(t)

    elements.append(Spacer(1, 6*mm))
    elements.append(Paragraph('GRL = Greek Restaurant Ltd (Paddington) | Pagema Ltd = Dark Kitchens (Brent, Chiswick, Peckham, Shoreditch, Wandsworth)', footnote_style))
    elements.append(Paragraph('Note: February 2026 Pagema expenses appear incomplete — likely pending Xero reconciliation.', footnote_style))

    doc.build(elements)
    print(f'  PDF: {path}')
    return path


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    print('Building consolidated P&L data...')
    data = build_consolidated()

    print('Generating outputs:')
    generate_html(data)
    generate_markdown(data)
    generate_excel(data)
    generate_pdf(data)

    print('\nDone! All files written to reports/output/')


if __name__ == '__main__':
    main()
