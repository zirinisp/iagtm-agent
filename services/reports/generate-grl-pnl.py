#!/usr/bin/env python3
"""
Generate GRL (Greek Restaurant Ltd) P&L reports in HTML, PDF, MD, and XLSX formats.
Reads structured JSON data from grl-pnl-paddington-dec25-feb26.json
"""
import json
import os

# --- Configuration ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE = os.path.join(BASE_DIR, 'grl-pnl-paddington-dec25-feb26.json')
OUTPUT_PREFIX = os.path.join(BASE_DIR, 'grl-pnl-paddington-dec25-feb26')

# Brand colors
NAVY = '#1B365D'
GOLD = '#D4A84B'
GREEN = '#2E7D32'
RED = '#C62828'
LIGHT_BG = '#F5F5F0'
ALT_ROW = '#EAE8E0'
WHITE = '#FFFFFF'

MONTH_KEYS = ['2025-12', '2026-01', '2026-02']
MONTH_LABELS = {'2025-12': 'Dec 2025', '2026-01': 'Jan 2026', '2026-02': 'Feb 2026'}
MONTH_LABELS_FULL = {'2025-12': 'December 2025', '2026-01': 'January 2026', '2026-02': 'February 2026'}

REPORT_TITLE = 'Greek Restaurant Ltd (Paddington) -- P&L (No Shareholder)'
DATE_RANGE = 'December 2025 -- February 2026'


def load_data():
    with open(INPUT_FILE, 'r') as f:
        return json.load(f)


def fmt_gbp(val):
    """Format as GBP with commas. Negatives in parentheses."""
    if val < 0:
        return f'({chr(163)}{abs(val):,.2f})'
    return f'{chr(163)}{val:,.2f}'


def fmt_pct(val):
    """Format as percentage to 1 decimal."""
    if val < 0:
        return f'({abs(val):.1f}%)'
    return f'{val:.1f}%'


def get_turnover(sections):
    """Get total income (turnover) for % calculations."""
    for s in sections:
        if s.get('total') and 'Income' in s['total'].get('account', ''):
            return s['total']['amount']
    return 0


def pct_of_turnover(amount, turnover):
    if turnover == 0:
        return 0
    return (amount / turnover) * 100


def flatten_sections(sections):
    """Flatten sections into a list of (label, amount, is_total, is_header, section_title) tuples."""
    rows = []
    for section in sections:
        title = section.get('title', '')
        if title:
            rows.append((title, None, False, True, title))
        for row in section.get('rows', []):
            rows.append((row['account'], row['amount'], False, False, title))
        if section.get('total'):
            t = section['total']
            rows.append((t['account'], t['amount'], True, False, title))
        # Special: untitled sections with a single row (Gross Profit, Net Profit)
        if not title and len(section.get('rows', [])) == 0 and section.get('total') is None:
            pass  # skip empty untitled
        if not title and section.get('rows'):
            for row in section['rows']:
                if row['account'] in ('Gross Profit', 'Net Profit', 'Operating Profit'):
                    # Already added above, just mark it
                    pass
    return rows


def flatten_for_display(sections):
    """Create display rows: list of dicts with label, amount, is_total, is_section_header, is_key_total."""
    rows = []
    key_totals = {'Gross Profit', 'Net Profit', 'Operating Profit',
                  'Total Income', 'Total Cost of Sales', 'Total Operating Expenses', 'Total Other Income'}

    for section in sections:
        title = section.get('title', '')

        # Section header
        if title:
            rows.append({
                'label': title, 'amount': None, 'is_total': False,
                'is_header': True, 'is_key_total': False
            })

        # Data rows
        for row in section.get('rows', []):
            is_key = row['account'] in key_totals
            rows.append({
                'label': row['account'], 'amount': row['amount'],
                'is_total': is_key, 'is_header': False, 'is_key_total': is_key
            })

        # Summary/total row
        if section.get('total'):
            t = section['total']
            rows.append({
                'label': t['account'], 'amount': t['amount'],
                'is_total': True, 'is_header': False,
                'is_key_total': t['account'] in key_totals
            })

    return rows


# ============================================================
# HTML Generation
# ============================================================
def generate_html(data):
    html_parts = [f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{REPORT_TITLE}</title>
<style>
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: {LIGHT_BG}; color: #333; padding: 20px; }}
  .container {{ max-width: 900px; margin: 0 auto; }}
  h1 {{ color: {NAVY}; font-size: 1.6em; margin-bottom: 4px; }}
  h2 {{ color: {NAVY}; font-size: 1.2em; margin: 30px 0 10px; border-bottom: 2px solid {GOLD}; padding-bottom: 6px; }}
  .subtitle {{ color: #666; font-size: 0.95em; margin-bottom: 20px; }}
  table {{ width: 100%; border-collapse: collapse; margin-bottom: 10px; font-size: 0.88em; }}
  th {{ background: {NAVY}; color: {WHITE}; padding: 8px 12px; text-align: left; font-weight: 600; }}
  th:nth-child(2), th:nth-child(3) {{ text-align: right; }}
  td {{ padding: 5px 12px; border-bottom: 1px solid #ddd; }}
  td:nth-child(2), td:nth-child(3) {{ text-align: right; font-variant-numeric: tabular-nums; }}
  tr:nth-child(even) td {{ background: {ALT_ROW}; }}
  tr.section-header td {{ background: {NAVY}; color: {WHITE}; font-weight: 600; padding: 6px 12px; }}
  tr.total-row td {{ font-weight: 700; border-top: 2px solid {NAVY}; border-bottom: 2px solid {NAVY}; background: #E8E6DE !important; }}
  tr.key-total td {{ font-weight: 700; border-top: 2px solid {GOLD}; border-bottom: 2px solid {GOLD}; background: #F0EDE0 !important; font-size: 1.02em; }}
  .negative {{ color: {RED}; }}
  .positive {{ color: {GREEN}; }}
  .footer {{ margin-top: 30px; font-size: 0.8em; color: #888; text-align: center; }}
</style>
</head>
<body>
<div class="container">
<h1>{REPORT_TITLE}</h1>
<p class="subtitle">{DATE_RANGE}</p>
"""]

    # Monthly tables
    for mk in MONTH_KEYS:
        month_data = data['months'][mk]
        sections = month_data['sections']
        turnover = get_turnover(sections)
        rows = flatten_for_display(sections)

        html_parts.append(f'<h2>{MONTH_LABELS_FULL[mk]}</h2>')
        html_parts.append('<table>')
        html_parts.append(f'<tr><th>Account</th><th>Paddington ({chr(163)})</th><th>% of Turnover</th></tr>')

        for row in rows:
            if row['is_header']:
                html_parts.append(f'<tr class="section-header"><td colspan="3">{row["label"]}</td></tr>')
            else:
                amt = row['amount'] if row['amount'] is not None else 0
                pct = pct_of_turnover(amt, turnover)
                css_class = ''
                if row['is_key_total']:
                    css_class = 'key-total'
                elif row['is_total']:
                    css_class = 'total-row'

                amt_str = fmt_gbp(amt)
                pct_str = fmt_pct(pct)
                color_class = 'negative' if amt < 0 else ''

                html_parts.append(
                    f'<tr class="{css_class}"><td>{"&nbsp;&nbsp;" if not row["is_total"] else ""}{row["label"]}</td>'
                    f'<td class="{color_class}">{amt_str}</td>'
                    f'<td class="{color_class}">{pct_str}</td></tr>'
                )

        html_parts.append('</table>')

    # Summary table
    html_parts.append('<h2>3-Month Summary (Dec 2025 - Feb 2026)</h2>')
    html_parts.append('<table>')
    html_parts.append(f'<tr><th>Metric</th>')
    for mk in MONTH_KEYS:
        html_parts.append(f'<th>{MONTH_LABELS[mk]}</th>')
    html_parts.append(f'<th>Total</th></tr>')

    summary_metrics = ['Total Income', 'Total Cost of Sales', 'Gross Profit',
                       'Total Other Income', 'Total Operating Expenses', 'Net Profit']

    for metric in summary_metrics:
        vals = []
        for mk in MONTH_KEYS:
            val = find_metric(data['months'][mk]['sections'], metric)
            vals.append(val)
        total = sum(vals)

        html_parts.append(f'<tr class="key-total"><td>{metric}</td>')
        for v in vals:
            color = 'negative' if v < 0 else ''
            html_parts.append(f'<td class="{color}">{fmt_gbp(v)}</td>')
        color = 'negative' if total < 0 else ''
        html_parts.append(f'<td class="{color}">{fmt_gbp(total)}</td></tr>')

    html_parts.append('</table>')
    html_parts.append(f'<div class="footer">Generated by IAGTM Agent | Data source: Xero API</div>')
    html_parts.append('</div></body></html>')

    with open(f'{OUTPUT_PREFIX}.html', 'w') as f:
        f.write('\n'.join(html_parts))
    print(f'HTML saved: {OUTPUT_PREFIX}.html')


def find_metric(sections, metric_name):
    """Find a metric value from sections by searching totals and rows."""
    for section in sections:
        if section.get('total') and section['total']['account'] == metric_name:
            return section['total']['amount']
        for row in section.get('rows', []):
            if row['account'] == metric_name:
                return row['amount']
    return 0


# ============================================================
# Markdown Generation
# ============================================================
def generate_md(data):
    lines = [f'# {REPORT_TITLE}', f'*{DATE_RANGE}*', '']

    for mk in MONTH_KEYS:
        month_data = data['months'][mk]
        sections = month_data['sections']
        turnover = get_turnover(sections)
        rows = flatten_for_display(sections)

        lines.append(f'## {MONTH_LABELS_FULL[mk]}')
        lines.append('')
        lines.append(f'| Account | Paddington ({chr(163)}) | % of Turnover |')
        lines.append('|---------|-------------------:|---------------:|')

        for row in rows:
            if row['is_header']:
                lines.append(f'| **{row["label"]}** | | |')
            else:
                amt = row['amount'] if row['amount'] is not None else 0
                pct = pct_of_turnover(amt, turnover)
                label = f'**{row["label"]}**' if row['is_total'] else f'  {row["label"]}'
                lines.append(f'| {label} | {fmt_gbp(amt)} | {fmt_pct(pct)} |')

        lines.append('')

    # Summary
    lines.append('## 3-Month Summary (Dec 2025 - Feb 2026)')
    lines.append('')
    header = f'| Metric | {" | ".join(MONTH_LABELS[mk] for mk in MONTH_KEYS)} | Total |'
    sep = '|--------|' + '|'.join(['---:'] * (len(MONTH_KEYS) + 1)) + '|'
    lines.append(header)
    lines.append(sep)

    summary_metrics = ['Total Income', 'Total Cost of Sales', 'Gross Profit',
                       'Total Other Income', 'Total Operating Expenses', 'Net Profit']

    for metric in summary_metrics:
        vals = [find_metric(data['months'][mk]['sections'], metric) for mk in MONTH_KEYS]
        total = sum(vals)
        row_vals = ' | '.join(fmt_gbp(v) for v in vals)
        lines.append(f'| **{metric}** | {row_vals} | {fmt_gbp(total)} |')

    lines.append('')
    lines.append('---')
    lines.append('*Generated by IAGTM Agent | Data source: Xero API*')

    with open(f'{OUTPUT_PREFIX}.md', 'w') as f:
        f.write('\n'.join(lines))
    print(f'MD saved: {OUTPUT_PREFIX}.md')


# ============================================================
# Excel Generation
# ============================================================
def generate_xlsx(data):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, numbers, Border, Side

    wb = Workbook()

    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1B365D', end_color='1B365D', fill_type='solid')
    gold_fill = PatternFill(start_color='D4A84B', end_color='D4A84B', fill_type='solid')
    total_fill = PatternFill(start_color='E8E6DE', end_color='E8E6DE', fill_type='solid')
    section_fill = PatternFill(start_color='1B365D', end_color='1B365D', fill_type='solid')
    red_font = Font(name='Calibri', color='C62828')
    bold_font = Font(name='Calibri', bold=True, size=11)
    normal_font = Font(name='Calibri', size=10)
    thin_border = Border(
        bottom=Side(style='thin', color='CCCCCC')
    )

    gbp_fmt = '#,##0.00;(#,##0.00)'
    pct_fmt = '0.0%'

    for idx, mk in enumerate(MONTH_KEYS):
        ws = wb.active if idx == 0 else wb.create_sheet()
        ws.title = MONTH_LABELS[mk]

        month_data = data['months'][mk]
        sections = month_data['sections']
        turnover = get_turnover(sections)
        rows = flatten_for_display(sections)

        # Title
        ws.merge_cells('A1:C1')
        ws['A1'] = f'{REPORT_TITLE} - {MONTH_LABELS_FULL[mk]}'
        ws['A1'].font = Font(name='Calibri', bold=True, size=14, color='1B365D')

        # Headers
        for col, header in enumerate(['Account', f'Paddington ({chr(163)})', '% of Turnover'], 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='right' if col > 1 else 'left')

        r = 4
        for row in rows:
            if row['is_header']:
                ws.cell(row=r, column=1, value=row['label']).font = Font(name='Calibri', bold=True, color='FFFFFF')
                for c in range(1, 4):
                    ws.cell(row=r, column=c).fill = section_fill
            else:
                amt = row['amount'] if row['amount'] is not None else 0
                pct = pct_of_turnover(amt, turnover) / 100  # Excel expects decimal

                label_cell = ws.cell(row=r, column=1, value=row['label'])
                amt_cell = ws.cell(row=r, column=2, value=amt)
                pct_cell = ws.cell(row=r, column=3, value=pct)

                amt_cell.number_format = gbp_fmt
                pct_cell.number_format = pct_fmt
                amt_cell.alignment = Alignment(horizontal='right')
                pct_cell.alignment = Alignment(horizontal='right')

                if row['is_total'] or row['is_key_total']:
                    for c in range(1, 4):
                        ws.cell(row=r, column=c).font = bold_font
                        ws.cell(row=r, column=c).fill = total_fill
                else:
                    label_cell.font = normal_font
                    label_cell.value = f'  {row["label"]}'
                    if amt < 0:
                        amt_cell.font = red_font
                        pct_cell.font = red_font

                for c in range(1, 4):
                    ws.cell(row=r, column=c).border = thin_border

            r += 1

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 18

    # Summary sheet
    ws = wb.create_sheet('Summary')
    ws.merge_cells('A1:E1')
    ws['A1'] = f'{REPORT_TITLE} - 3-Month Summary'
    ws['A1'].font = Font(name='Calibri', bold=True, size=14, color='1B365D')

    headers = ['Metric'] + [MONTH_LABELS[mk] for mk in MONTH_KEYS] + ['Total']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='right' if col > 1 else 'left')

    summary_metrics = ['Total Income', 'Total Cost of Sales', 'Gross Profit',
                       'Total Other Income', 'Total Operating Expenses', 'Net Profit']

    for i, metric in enumerate(summary_metrics, 4):
        ws.cell(row=i, column=1, value=metric).font = bold_font
        vals = [find_metric(data['months'][mk]['sections'], metric) for mk in MONTH_KEYS]
        total = sum(vals)
        for j, v in enumerate(vals, 2):
            cell = ws.cell(row=i, column=j, value=v)
            cell.number_format = gbp_fmt
            cell.alignment = Alignment(horizontal='right')
            if v < 0:
                cell.font = red_font
        total_cell = ws.cell(row=i, column=len(MONTH_KEYS) + 2, value=total)
        total_cell.number_format = gbp_fmt
        total_cell.alignment = Alignment(horizontal='right')
        total_cell.font = bold_font
        if total < 0:
            total_cell.font = Font(name='Calibri', bold=True, color='C62828')

        for c in range(1, len(headers) + 1):
            ws.cell(row=i, column=c).fill = total_fill
            ws.cell(row=i, column=c).border = thin_border

    ws.column_dimensions['A'].width = 30
    for c in ['B', 'C', 'D', 'E']:
        ws.column_dimensions[c].width = 18

    wb.save(f'{OUTPUT_PREFIX}.xlsx')
    print(f'XLSX saved: {OUTPUT_PREFIX}.xlsx')


# ============================================================
# PDF Generation
# ============================================================
def generate_pdf(data):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    doc = SimpleDocTemplate(f'{OUTPUT_PREFIX}.pdf', pagesize=A4,
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=15*mm, bottomMargin=15*mm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title2', parent=styles['Title'],
                                  textColor=colors.HexColor(NAVY), fontSize=16, spaceAfter=4)
    subtitle_style = ParagraphStyle('Subtitle2', parent=styles['Normal'],
                                     textColor=colors.HexColor('#666666'), fontSize=10, spaceAfter=16)
    section_style = ParagraphStyle('Section2', parent=styles['Heading2'],
                                    textColor=colors.HexColor(NAVY), fontSize=13, spaceAfter=6,
                                    spaceBefore=16)

    navy_color = colors.HexColor(NAVY)
    gold_color = colors.HexColor(GOLD)
    red_color = colors.HexColor(RED)
    alt_color = colors.HexColor(ALT_ROW)
    total_bg = colors.HexColor('#E8E6DE')

    elements = []
    elements.append(Paragraph(REPORT_TITLE, title_style))
    elements.append(Paragraph(DATE_RANGE, subtitle_style))

    col_widths = [110*mm, 35*mm, 30*mm]

    for mk in MONTH_KEYS:
        month_data = data['months'][mk]
        sections = month_data['sections']
        turnover = get_turnover(sections)
        rows = flatten_for_display(sections)

        elements.append(Paragraph(MONTH_LABELS_FULL[mk], section_style))

        table_data = [[
            Paragraph(f'<b>Account</b>', styles['Normal']),
            Paragraph(f'<b>Paddington ({chr(163)})</b>', styles['Normal']),
            Paragraph(f'<b>% of Turnover</b>', styles['Normal']),
        ]]

        style_commands = [
            ('BACKGROUND', (0, 0), (-1, 0), navy_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ]

        row_idx = 1
        for row in rows:
            if row['is_header']:
                table_data.append([row['label'], '', ''])
                style_commands.append(('BACKGROUND', (0, row_idx), (-1, row_idx), navy_color))
                style_commands.append(('TEXTCOLOR', (0, row_idx), (-1, row_idx), colors.white))
                style_commands.append(('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold'))
                style_commands.append(('SPAN', (0, row_idx), (-1, row_idx)))
            else:
                amt = row['amount'] if row['amount'] is not None else 0
                pct = pct_of_turnover(amt, turnover)
                label = row['label']
                if not row['is_total']:
                    label = f'  {label}'

                table_data.append([label, fmt_gbp(amt), fmt_pct(pct)])

                if row['is_total'] or row['is_key_total']:
                    style_commands.append(('BACKGROUND', (0, row_idx), (-1, row_idx), total_bg))
                    style_commands.append(('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold'))
                elif row_idx % 2 == 0:
                    style_commands.append(('BACKGROUND', (0, row_idx), (-1, row_idx), alt_color))

                if amt < 0:
                    style_commands.append(('TEXTCOLOR', (1, row_idx), (2, row_idx), red_color))

            row_idx += 1

        t = Table(table_data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle(style_commands))
        elements.append(t)
        elements.append(Spacer(1, 6*mm))

    # Summary
    elements.append(Paragraph('3-Month Summary (Dec 2025 - Feb 2026)', section_style))

    summary_metrics = ['Total Income', 'Total Cost of Sales', 'Gross Profit',
                       'Total Other Income', 'Total Operating Expenses', 'Net Profit']

    sum_col_widths = [55*mm, 30*mm, 30*mm, 30*mm, 30*mm]
    sum_data = [['Metric'] + [MONTH_LABELS[mk] for mk in MONTH_KEYS] + ['Total']]
    sum_styles = [
        ('BACKGROUND', (0, 0), (-1, 0), navy_color),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
    ]

    for i, metric in enumerate(summary_metrics):
        vals = [find_metric(data['months'][mk]['sections'], metric) for mk in MONTH_KEYS]
        total = sum(vals)
        sum_data.append([metric] + [fmt_gbp(v) for v in vals] + [fmt_gbp(total)])
        ri = i + 1
        sum_styles.append(('BACKGROUND', (0, ri), (-1, ri), total_bg))
        sum_styles.append(('FONTNAME', (0, ri), (-1, ri), 'Helvetica-Bold'))
        for j, v in enumerate(vals + [total]):
            if v < 0:
                sum_styles.append(('TEXTCOLOR', (j + 1, ri), (j + 1, ri), red_color))

    t = Table(sum_data, colWidths=sum_col_widths, repeatRows=1)
    t.setStyle(TableStyle(sum_styles))
    elements.append(t)

    # Footer
    elements.append(Spacer(1, 10*mm))
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'],
                                   textColor=colors.HexColor('#888888'), fontSize=7,
                                   alignment=1)
    elements.append(Paragraph('Generated by IAGTM Agent | Data source: Xero API', footer_style))

    doc.build(elements)
    print(f'PDF saved: {OUTPUT_PREFIX}.pdf')


# ============================================================
# Main
# ============================================================
if __name__ == '__main__':
    data = load_data()
    generate_html(data)
    generate_md(data)
    generate_xlsx(data)
    generate_pdf(data)
    print('\nAll reports generated successfully.')
